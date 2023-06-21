import openpyxl


def read_schedule(input_file, group_name):
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    # Получаем количество строк и столбцов в таблице
    rows = sheet.max_row
    cols = sheet.max_column

    # Находим столбец с указанной группой
    group_col = None
    for col in range(2, cols + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value == group_name:
            group_col = col
            break

    if group_col is None:
        raise ValueError(f"Группа {group_name} не найдена в таблице.")

    # Читаем заголовки колонок с группами и уроками
    groups = []
    lessons = []
    for col in range(group_col, cols + 1):
        group = sheet.cell(row=1, column=col).value
        groups.append(group)
        lesson = sheet.cell(row=2, column=col).value
        lessons.append(lesson)

    # Читаем дни недели и уроки для каждой группы, начиная с указанной группы
    schedule_data = {}
    for row in range(3, rows + 1):
        day = sheet.cell(row=row, column=1).value
        schedule_data[day] = {}
        for col in range(group_col, cols + 1):
            group = groups[col - group_col]
            lesson = sheet.cell(row=row, column=col).value

            # Если уже есть урок для данной группы в этот день, добавляем его к предыдущему
            if group in schedule_data[day]:
                if lesson is not None:
                    lesson += "\n" + schedule_data[day][group]
            else:
                if lesson is None:
                    lesson = ""

            schedule_data[day][group] = lesson

    return groups, lessons, schedule_data


def print_column_by_group(groups, lessons, group_name):
    group_index = groups.index(group_name)
    lesson = lessons[group_index]
